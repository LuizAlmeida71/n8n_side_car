from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import JSONResponse
from openpyxl import Workbook, load_workbook
import tempfile
import fitz  # PyMuPDF
import base64
import os
from fpdf import FPDF
import traceback
from collections import defaultdict
import re
from datetime import datetime, timedelta
from typing import List

app = FastAPI()

# --- DEFINIÇÕES GLOBAIS E FUNÇÕES AUXILIARES ---

FONT_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"

MONTH_MAP = {
    'JANEIRO': 1, 'FEVEREIRO': 2, 'MARÇO': 3, 'ABRIL': 4, 'MAIO': 5,
    'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9, 'OUTUBRO': 10,
    'NOVEMBRO': 11, 'DEZEMBRO': 12
}

HORARIOS_TURNO = {
    "MANHÃ": {"inicio": "07:00", "fim": "13:00"},
    "TARDE": {"inicio": "13:00", "fim": "19:00"},
    "NOITE (início)": {"inicio": "19:00", "fim": "01:00"},
    "NOITE (fim)": {"inicio": "01:00", "fim": "07:00"},
}

def parse_mes_ano_pacs(text: str):
    month_regex = '|'.join(MONTH_MAP.keys())
    match = re.search(
        r'(?:MÊS\s*(?:DE)?\s*)?(' + month_regex + r')\s*(?:DE\s*|[/|-]?)\s*(\d{4})',
        text.upper(),
        re.IGNORECASE
    )
    if not match:
        return None, None
    mes_nome, ano_str = match.groups()
    mes = MONTH_MAP.get(mes_nome.upper())
    ano = int(ano_str)
    return mes, ano

def interpretar_turno_pacs(token: str, medico_setor: str):
    if not token or not isinstance(token, str):
        return []
    token_clean = token.replace('\n', ' ').replace('/', ' ').replace(' ', '')
    tokens = list(token_clean)
    turnos_finais = []
    if "TOTAL" in token.upper() or "PL" in token.upper():
        return []
    for t in tokens:
        if t == 'M': turnos_finais.append({"turno": "MANHÃ"})
        elif t == 'T': turnos_finais.append({"turno": "TARDE"})
        elif t == 'D':
            turnos_finais.append({"turno": "MANHÃ"})
            turnos_finais.append({"turno": "TARDE"})
        elif t == 'N':
            turnos_finais.append({"turno": "NOITE (início)"})
            turnos_finais.append({"turno": "NOITE (fim)"})
        elif t == 'n': turnos_finais.append({"turno": "NOITE (início)"})
    return turnos_finais

def is_valid_professional_name(name: str):
    if not name or not isinstance(name, str): return False
    name_upper = name.strip().upper()
    ignored_keywords = ["NOME COMPLETO", "LEGENDA", "ASSINATURA", "ASSINADO", "COMPLETO", "CARGO", "MATRÍCULA"]
    return not any(keyword in name_upper for keyword in ignored_keywords) and \
           (len(name.split()) >= 2 or name.isupper())

def dedup_plantao(lista_plantoes: list):
    seen = set()
    result = []
    for p in lista_plantoes:
        key = (p["data"], p["turno"], p["inicio"], p["fim"])
        if key not in seen:
            seen.add(key)
            result.append(p)
    return result

def extrair_metadados_pagina(page_text):
    unidade = re.search(r'UNIDADE[:\s-]*(.+?)(UNIDADE|SETOR|MÊS|ESCALA|$)', page_text.replace('\n', ' '), re.I)
    setor = re.search(r'UNIDADE[/\s-]*SETOR[:\s-]*(.+?)(MÊS|ESCALA|$)', page_text.replace('\n', ' '), re.I)
    return unidade.group(1).strip() if unidade else None, setor.group(1).strip() if setor else None

def parse_mes_ano_pdf(text):
    match = re.search(r'MÊS[\s/:]*([A-ZÇÃ]+)[\s/]*(\d{4})', text.upper())
    if not match: return None, None
    mes_nome, ano_str = match.groups()
    return MONTH_MAP.get(mes_nome), int(ano_str)

def interpretar_turno_pdf(token):
    token = token.replace('\n', '').replace('/', '').replace(' ', '').upper()
    turnos_finais = []
    for t in token:
        if t == 'M': turnos_finais.append("MANHÃ")
        elif t == 'T': turnos_finais.append("TARDE")
        elif t == 'D': turnos_finais.extend(["MANHÃ", "TARDE"])
        elif t == 'N': turnos_finais.extend(["NOITE (início)", "NOITE (fim)"])
        elif t == 'n': turnos_finais.append("NOITE (início)")
    return turnos_finais

# --- ENDPOINTS DA API ---

@app.post("/xlsx-to-json")
async def convert_xlsx_to_json(file: UploadFile = File(...)):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            contents = await file.read()
            tmp.write(contents)
            tmp_path = tmp.name

        workbook = load_workbook(filename=tmp_path, data_only=True)
        all_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows: continue

            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            data_rows = rows[1:]
            data = []

            for row in data_rows:
                row_dict = {headers[i]: row[i] for i in range(len(headers)) if headers[i] != ""}
                data.append(row_dict)

            all_data[sheet_name] = data

        return JSONResponse(content=all_data)
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.post("/split-pdf")
async def split_pdf(file: UploadFile = File(...)):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            contents = await file.read()
            tmp.write(contents)
            tmp_path = tmp.name

        pages_b64 = []
        with fitz.open(tmp_path) as doc:
            for i in range(len(doc)):
                single_page = fitz.open()
                single_page.insert_pdf(doc, from_page=i, to_page=i)
                page_path = f"/tmp/page_{i+1}.pdf"
                single_page.save(page_path, garbage=4, deflate=True, incremental=False)

                with open(page_path, "rb") as f:
                    b64_content = base64.b64encode(f.read()).decode("utf-8")
                    pages_b64.append({
                        "page": i + 1,
                        "file_base64": b64_content,
                        "filename": f"page_{i+1}.pdf"
                    })
                os.remove(page_path)

        os.remove(tmp_path)
        return JSONResponse(content={"pages": pages_b64})
    except Exception as e:
        if 'tmp_path' in locals() and os.path.exists(tmp_path):
            os.remove(tmp_path)
        return JSONResponse(content={"error": str(e), "trace": traceback.format_exc()}, status_code=500)

@app.post("/split-pdf-base64")
async def split_pdf_base64(request: Request):
    try:
        body = await request.json()
        b64 = body.get("base64")
        if not b64:
            return JSONResponse(content={"error": "Campo 'base64' ausente"}, status_code=400)

        pdf_bytes = base64.b64decode(b64)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        pages_b64 = []

        for i in range(len(doc)):
            single_page = fitz.open()
            single_page.insert_pdf(doc, from_page=i, to_page=i)
            page_path = f"/tmp/page_{i+1}.pdf"
            single_page.save(page_path, garbage=4, deflate=True, incremental=False)
            with open(page_path, "rb") as f:
                b64_content = base64.b64encode(f.read()).decode("utf-8")
                pages_b64.append({
                    "page": i + 1,
                    "file_base64": b64_content,
                    "filename": f"page_{i+1}.pdf"
                })
            os.remove(page_path)

        return JSONResponse(content={"pages": pages_b64})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

# --- INÍCIO normaliza-escala-from-pdf ---
@app.post("/normaliza-escala-from-pdf")
async def normaliza_escala_from_pdf(request: Request):
    try:
        body = await request.json()
        pages = body["pages"] if isinstance(body, dict) else body
        all_table_rows, last_unidade, last_setor, last_mes, last_ano = [], None, None, None, None

        for page_data in pages:
            b64_data = page_data.get("file_base64") or page_data.get("bae64") or page_data.get("base64")
            if not b64_data: continue

            pdf_bytes = base64.b64decode(b64_data)
            with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                page_text = doc[0].get_text("text")
                unidade, setor = extrair_metadados_pagina(page_text)
                if unidade: last_unidade = unidade
                if setor: last_setor = setor
                mes, ano = parse_mes_ano_pdf(page_text)
                if mes: last_mes = mes
                if ano: last_ano = ano

                tabelas = [t.extract() for t in doc[0].find_tables() if t.extract()]
                if tabelas:
                    tabela = tabelas[0]
                    all_table_rows.extend(tabela)

        header_row_idx = next((i for i, row in enumerate(all_table_rows) if row and "NOME COMPLETO" in ''.join([str(cell or '') for cell in row]).upper()), None)
        if header_row_idx is None:
            return JSONResponse({"error": "Cabeçalho não encontrado."}, status_code=400)

        dias_row = all_table_rows[header_row_idx + 1]
        header_map = {idx: col.strip() for idx, col in enumerate(dias_row) if isinstance(col, (str, int)) and str(col).strip().isdigit()}
        nome_idx = next(idx for idx, col in enumerate(all_table_rows[header_row_idx]) if "NOME COMPLETO" in str(col).upper())

        profissionais_data = defaultdict(lambda: defaultdict(list))
        for row in all_table_rows[header_row_idx+2:]:
            nome_bruto = row[nome_idx]
            if not is_valid_professional_name(nome_bruto): continue
            nome = ' '.join(nome_bruto.split())

            for idx, dia in header_map.items():
                if idx < len(row) and row[idx]:
                    turnos = interpretar_turno_pdf(row[idx])
                    for turno in turnos:
                        horarios = HORARIOS_TURNO.get(turno)
                        dia_plantao = datetime(last_ano, last_mes, int(dia))
                        if turno == "NOITE (fim)": dia_plantao += timedelta(days=1)
                        profissionais_data[nome]["plantoes"].append({
                            "dia": dia_plantao.day,
                            "data": dia_plantao.strftime('%d/%m/%Y'),
                            "turno": turno,
                            "inicio": horarios["inicio"],
                            "fim": horarios["fim"]
                        })

        lista_profissionais_final = [{
            "medico_nome": nome,
            "medico_setor": last_setor or "NÃO INFORMADO",
            "plantoes": sorted(plantoes["plantoes"], key=lambda x: (datetime.strptime(x["data"], '%d/%m/%Y'), x["inicio"]))
        } for nome, plantoes in profissionais_data.items()]

        mes_nome_str = [k for k, v in MONTH_MAP.items() if v == last_mes][0]
        return JSONResponse(content=[{
            "unidade_escala": last_unidade or "NÃO INFORMADO",
            "mes_ano_escala": f"{mes_nome_str}/{last_ano}",
            "profissionais": lista_profissionais_final
        }])
    except Exception as e:
        return JSONResponse({"error": str(e), "trace": traceback.format_exc()}, status_code=500)
# --- FIM normaliza-escala-from-pdf ---


@app.post("/text-to-pdf")
async def text_to_pdf(request: Request):
    try:
        data = await request.json()
        raw_text = data.get("text", "")
        filename = data.get("filename", "saida.pdf")

        if not os.path.exists(FONT_PATH):
            raise RuntimeError(f"Fonte não encontrada em: {FONT_PATH}")

        clean_text = raw_text.replace("\r", "").replace("\n", " ")
        clean_text = " ".join(clean_text.split())
        lines = [clean_text[i:i+120] for i in range(0, len(clean_text), 120)]

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_font("DejaVu", "", FONT_PATH, uni=True)
        pdf.set_font("DejaVu", size=10)

        for line in lines:
            pdf.multi_cell(w=190, h=8, txt=line)
        
        pdf_output = pdf.output(dest='S')
        pdf_bytes = pdf_output.encode('latin1')
        base64_pdf = base64.b64encode(pdf_bytes).decode("utf-8")

        return JSONResponse(content={"file_base64": base64_pdf, "filename": filename})
    except Exception as e:
        return JSONResponse(content={"error": str(e), "trace": traceback.format_exc()}, status_code=500)


# --- INÍCIO normaliza-escala-PACS ---
@app.post("/normaliza-escala-PACS")
async def normaliza_escala_PACS(request: Request):
    try:
        body = await request.json()
        full_text = ""
        all_table_rows = []
        last_unidade, last_setor = None, None
        last_mes, last_ano = None, None

        for page_data in body:
            b64_data = page_data.get("base64") or page_data.get("bae64")
            if not b64_data:
                continue
            pdf_bytes = base64.b64decode(b64_data)
            with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                page = doc[0]
                page_text = page.get_text("text")
                if not full_text:
                    full_text = page_text
                for table in page.find_tables():
                    extracted = table.extract()
                    if extracted:
                        all_table_rows.extend(extracted)

        unidade_match = re.search(r'UNIDADE:\s*(.*?)\n', full_text, re.IGNORECASE)
        setor_match = re.search(r'SETOR:\s*(.*?)\n', full_text, re.IGNORECASE)
        mes, ano = parse_mes_ano_pacs(full_text)

        last_unidade = unidade_match.group(1).strip() if unidade_match else "NÃO INFORMADO"
        last_setor = setor_match.group(1).strip() if setor_match else "NÃO INFORMADO"
        last_mes, last_ano = mes, ano

        if last_mes is None or last_ano is None:
            return JSONResponse(content={"error": "Mês/Ano não encontrados."}, status_code=400)

        profissionais_data = defaultdict(lambda: {"info_rows": []})
        header_map = None
        nome_idx = None
        idx_linha = 0
        last_name = None

        while idx_linha < len(all_table_rows):
            row = all_table_rows[idx_linha]
            if not row or not any(row):
                idx_linha += 1
                continue

            if any("NOME" in str(cell or '').upper() and "COMPLETO" in str(cell or '').upper() for cell in row):
                first_col_is_index = str(row[0]).strip().isdigit()
                start_offset = 1 if first_col_is_index else 0
                header_row = row[start_offset:]
                header_map = {}
                for i, col_name in enumerate(header_row):
                    clean_name_upper = str(col_name or '').replace('\n', ' ').strip().upper()
                    col_pos = i + start_offset

                    if "NOME COMPLETO" in clean_name_upper: header_map["NOME COMPLETO"] = col_pos
                    elif "CARGO" in clean_name_upper: header_map["CARGO"] = col_pos
                    elif "VÍNCULO" in clean_name_upper or "VINCULO" in clean_name_upper: header_map["VÍNCULO"] = col_pos
                    elif "CONSELHO" in clean_name_upper or "CRM" in clean_name_upper: header_map["CRM"] = col_pos
                    else:
                        day_match = re.match(r'^(\d{1,2})(?:\D|$)', str(col_name or '').strip())
                        if day_match:
                            day_number = int(day_match.group(1))
                            if 1 <= day_number <= 31:
                                header_map[day_number] = col_pos

                nome_idx = header_map.get("NOME COMPLETO")
                last_name = None
                idx_linha += 1
                continue

            if not header_map or nome_idx is None:
                idx_linha += 1
                continue

            nome_bruto = row[nome_idx] if nome_idx < len(row) else None
            if nome_bruto and is_valid_professional_name(nome_bruto):
                last_name = nome_bruto.replace('\n', ' ').strip()
            elif nome_bruto and last_name and len(nome_bruto.strip().split()) == 1:
                last_name = f"{last_name} {nome_bruto.strip()}"

            if last_name:
                new_row = list(row)
                if nome_idx < len(new_row): new_row[nome_idx] = last_name
                profissionais_data[last_name]["info_rows"].append(new_row)

            idx_linha += 1

        lista_profissionais_final = []
        for nome, data in profissionais_data.items():
            info_rows = data["info_rows"]
            if not info_rows: continue

            primeira_linha = info_rows[0]

            def get_cell_value(col_name, default="N/I"):
                idx = header_map.get(col_name)
                if idx is not None and idx < len(primeira_linha) and primeira_linha[idx]:
                    return str(primeira_linha[idx]).strip()
                return default

            profissional_obj = {
                "medico_nome": nome.replace('\n', ' ').strip(),
                "medico_crm": get_cell_value("CRM").replace('\n', ' ').strip(),
                "medico_especialidade": get_cell_value("CARGO").replace('\n', ' ').strip(),
                "medico_vinculo": get_cell_value("VÍNCULO").replace('\n', ' ').strip(),
                "medico_setor": last_setor,
                "medico_unidade": last_unidade,
                "plantoes": []
            }

            if "PAES" not in profissional_obj["medico_vinculo"].upper():
                continue

            plantoes_brutos = defaultdict(list)
            for row in info_rows:
                for dia, col_idx in header_map.items():
                    if isinstance(dia, int) and col_idx < len(row) and row[col_idx]:
                        plantoes_brutos[dia].append(str(row[col_idx]).strip())

            for dia, tokens in sorted(plantoes_brutos.items()):
                for token in tokens:
                    turnos = interpretar_turno_pacs(token, last_setor)
                    try:
                        data_plantao = datetime(last_ano, last_mes, dia)
                    except ValueError:
                        continue

                    for turno_info in turnos:
                        horarios = HORARIOS_TURNO.get(turno_info["turno"], {})
                        data_inicio = data_plantao
                        if turno_info["turno"] == "NOITE (fim)":
                            data_inicio += timedelta(days=1)
                        profissional_obj["plantoes"].append({
                            "data": data_inicio.strftime('%d/%m/%Y'),
                            "dia": data_inicio.day,
                            "turno": turno_info["turno"],
                            "setor": last_setor,
                            "inicio": horarios.get("inicio"),
                            "fim": horarios.get("fim")
                        })

            profissional_obj["plantoes"] = dedup_plantao(profissional_obj["plantoes"])
            if profissional_obj["plantoes"]:
                profissional_obj["plantoes"].sort(key=lambda p: (datetime.strptime(p['data'], '%d/%m/%Y').toordinal(), p["inicio"] or ""))
                lista_profissionais_final.append(profissional_obj)

        lista_profissionais_final.sort(key=lambda p: p['medico_nome'])
        mes_nome_str = list(MONTH_MAP.keys())[list(MONTH_MAP.values()).index(last_mes)]
        final_output = [{"unidade_escala": last_unidade, "mes_ano_escala": f"{mes_nome_str}/{last_ano}", "profissionais": lista_profissionais_final}]

        return JSONResponse(content=final_output)

    except Exception as e:
        return JSONResponse(content={"error": str(e), "trace": traceback.format_exc()}, status_code=500)
# --- FIM normaliza-escala-PACS ---




# --- INICIO normaliza-MATERNIDADE-MATRICIAL ---
MONTH_MAP = {
    'JANEIRO': 1, 'FEVEREIRO': 2, 'MARÇO': 3, 'ABRIL': 4, 'MAIO': 5,
    'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9, 'OUTUBRO': 10,
    'NOVEMBRO': 11, 'DEZEMBRO': 12
}

HORARIOS_TURNO = {
    "MANHÃ": {"inicio": "07:00", "fim": "13:00"},
    "TARDE": {"inicio": "13:00", "fim": "19:00"},
    "NOITE (início)": {"inicio": "19:00", "fim": "01:00"},
    "NOITE (fim)": {"inicio": "01:00", "fim": "07:00"},
}

# Lista de profissionais da RP PAES como âncora
PROFISSIONAIS_ANCHOR = [
    {"medico_nome": "MARCO ANTÔNIO LEAL SANTOS", "medico_setor": "CAMED/BLOCOS/ISOLAMENTO/UTIN/UCINco/UCINca", "medico_unidade": "HMINSN"},
    {"medico_nome": "MOACIR BARBOSA NETO", "medico_setor": "CAMED/BLOCOS/ISOLAMENTO/UTIN/UCINco/UCINca", "medico_unidade": "HMINSN"},
    {"medico_nome": "ROBERTO ANDRADE LIMA", "medico_setor": "CAMED/BLOCOS/ISOLAMENTO/UTIN/UCINco/UCINca", "medico_unidade": "HMINSN"},
    {"medico_nome": "MARYCASSIELY RODRIGUES TIZOLIM", "medico_setor": "NIR/ISOLAMENTO/BLOCOS/UTIN/UTIM", "medico_unidade": "HMINSN"},
    {"medico_nome": "CIBELE LOUSANE PINHO MOTA", "medico_setor": "NIR/ISOLAMENTO/BLOCOS/UTIN/UTIM", "medico_unidade": "HMINSN"},
    {"medico_nome": "MANOEL MESSIAS DOS SANTOS NETO", "medico_setor": "NIR/ISOLAMENTO/BLOCOS/UTIN/UTIM", "medico_unidade": "HMINSN"}
]

def parse_mes_ano(text):
    month_regex = '|'.join(MONTH_MAP.keys())
    match = re.search(r'(?:MÊS[^A-Z]*)?(' + month_regex + r')[^\d]*(\d{4})', text.upper())
    if not match:
        return None, None
    mes_nome, ano_str = match.groups()
    return MONTH_MAP.get(mes_nome.upper()), int(ano_str)

def extrair_setor_e_unidade(text, lines, table_data=None):
    """
    Extrai setor e unidade do texto com robustez aprimorada.
    """
    text_normalized = text.upper().replace('Ç', 'C').replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
    nome_unidade = "NÃO INFORMADO"
    nome_setor = "NÃO INFORMADO"

    # Mapeamento de abreviações
    UNIDADE_ABREVIACOES = {
        "HMINSN": "HOSPITAL MATERNO INFANTIL NOSSA SENHORA DE NAZARETH"
    }

    # Padrões de extração
    pattern_unidade = r'UNIDADE:\s*([^\n]*(?:\n\s*[^\n]*)*?)'
    pattern_setor = r'UNIDADE/SETOR:\s*([^(\n]*(?:\n\s*[^(\n]*)*?)(?=\s*(ESCALA\s+DE\s+(SERVIÇO|SERVICO):|\n\s*NOME|\Z))'

    # Depuração das linhas
    print(f"Linhas processadas: {lines}")

    # Verificar todas as linhas
    for i, line in enumerate(lines):
        line_normalized = line.upper().replace('Ç', 'C').replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
        print(f"Linha {i}: {line_normalized}")
        unidade_match = re.search(pattern_unidade, line_normalized, re.IGNORECASE)
        if unidade_match:
            nome_unidade = unidade_match.group(1).strip()
            nome_unidade = UNIDADE_ABREVIACOES.get(nome_unidade, nome_unidade)
        setor_match = re.search(pattern_setor, line_normalized, re.IGNORECASE)
        if setor_match:
            nome_setor = setor_match.group(1).strip()
            nome_setor = re.sub(r'\s*(ESCALA\s+DE\s+(SERVIÇO|SERVICO):.*|\Z)', '', nome_setor).strip()
        if nome_setor != "NÃO INFORMADO" and nome_unidade != "NÃO INFORMADO":
            break

    # Fallback para texto completo
    if nome_unidade == "NÃO INFORMADO":
        unidade_match = re.search(pattern_unidade, text_normalized, re.IGNORECASE | re.DOTALL)
        if unidade_match:
            nome_unidade = unidade_match.group(1).strip()
            nome_unidade = UNIDADE_ABREVIACOES.get(nome_unidade, nome_unidade)
    if nome_setor == "NÃO INFORMADO":
        setor_match = re.search(pattern_setor, text_normalized, re.IGNORECASE | re.DOTALL)
        if setor_match:
            nome_setor = setor_match.group(1).strip()
            nome_setor = re.sub(r'\s*(ESCALA\s+DE\s+(SERVIÇO|SERVICO):.*|\Z)', '', nome_setor).strip()

    # Fallback aprimorado: Inferir a partir de abreviações ou tabela
    if nome_unidade == "NÃO INFORMADO" and any(abrev in text_normalized for abrev in UNIDADE_ABREVIACOES):
        for abrev, unidade in UNIDADE_ABREVIACOES.items():
            if abrev in text_normalized:
                nome_unidade = unidade
                break
    if nome_setor == "NÃO INFORMADO" and table_data and len(table_data) > 0:
        header_text = " ".join(str(cell or "").strip().upper() for cell in table_data[0][:5] if cell)  # Ampliar para 5 colunas
        if "CENTRO OBSTETRICO" in header_text or "OBSTETRICIA" in header_text or "GINECOLOGIA" in header_text:
            nome_setor = "CENTRO OBSTETRICO/ ORQUIDEAS"
        elif "TRIAGEM" in header_text:
            nome_setor = "TRIAGEM"
        elif "CAMED" in header_text:
            nome_setor = "CAMED/BLOCOS/ISOLAMENTO/UTIN/UCINco/UCINca"
        elif "ISOLAMENTO" in header_text:
            nome_setor = "CAMED/BLOCOS/ISOLAMENTO/UTIN/UCINco/UCINca"

    print(f"Extraído - Unidade: {nome_unidade}, Setor: {nome_setor} (Texto extraído: {text[:500]}...)")  # Mais contexto
    return nome_unidade, nome_setor

def interpretar_turno(token):
    if not token or not isinstance(token, str):
        return []
    token_clean = token.replace('\n', '').replace(' ', '').replace('/', '')
    if "TOTAL" in token.upper() or "PL" in token.upper():
        return []
    if len(token_clean) >= 2 and token_clean[-1].upper() in ['M', 'T', 'D', 'N']:
        tokens = [token_clean[-1].upper()]
    else:
        tokens = list(token_clean.upper())
    turnos = []
    for t in tokens:
        if t == 'M': turnos.append({"turno": "MANHÃ"})
        elif t == 'T': turnos.append({"turno": "TARDE"})
        elif t == 'D': turnos.append({"turno": "MANHÃ"}); turnos.append({"turno": "TARDE"})
        elif t == 'N': turnos.append({"turno": "NOITE (início)"}); turnos.append({"turno": "NOITE (fim)"})
    return turnos

def dedup_plantao(plantoes):
    seen = set()
    result = []
    for p in plantoes:
        key = (p["data"], p["turno"], p["inicio"], p["fim"])
        if key not in seen:
            seen.add(key)
            result.append(p)
    return result

def processar_pagina_pdf(b64_content, page_info=""):
    try:
        pdf_bytes = base64.b64decode(b64_content)
        profissionais = []

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                text = page.extract_text() or ""
                lines = text.splitlines()
                lines = [l for l in lines if not l.strip().startswith("Governo do Estado")]
                text = '\n'.join(lines)

                # Extrair unidade e setor
                nome_unidade, nome_setor = extrair_setor_e_unidade(text, lines, tables[0] if tables else None)
                if nome_setor == "NÃO INFORMADO":
                    print(f"AVISO: Setor não encontrado em {page_info}, página {page_num + 1}")
                if nome_unidade == "NÃO INFORMADO":
                    print(f"AVISO: Unidade não encontrada em {page_info}, página {page_num + 1}")

                # Extrair mês e ano
                mes, ano = parse_mes_ano(text)
                if not mes or not ano:
                    print(f"AVISO: Mês/Ano não encontrado em {page_info}, página {page_num + 1}")
                    continue

                print(f"Processando {page_info}, página {page_num + 1}: Unidade={nome_unidade}, Setor={nome_setor}, Mês={mes}/{ano}")

                # Processar tabelas
                if tables:
                    for table_num, table in enumerate(tables):
                        header = {}
                        for row_idx, row in enumerate(table):
                            if not header and any("NOME" in str(c).upper() for c in row if c):
                                for i, col in enumerate(row):
                                    col_clean = (col or "").strip().upper()
                                    if "NOME" in col_clean: header["nome"] = i
                                    elif "CARGO" in col_clean: header["cargo"] = i
                                    elif "VÍNCULO" in col_clean or "VINCULO" in col_clean: header["vinculo"] = i
                                    elif "CRM" in col_clean or "CONSELHO" in col_clean: header["crm"] = i
                                    elif "MATRÍCULA" in col_clean or "MATRICULA" in col_clean: header["matricula"] = i
                                    elif re.fullmatch(r"\d{1,2}", col_clean): header[int(col_clean)] = i
                                continue
                            if "nome" not in header or not row or not row[header["nome"]]:
                                continue
                            nome = str(row[header["nome"]]).replace('\n', ' ').strip()
                            if "SERVIDOR QUE ESTA FORA DA ESCALA" in nome.upper():
                                break
                            crm = str(row[header.get("crm", -1)] or "").replace('\n', ' ').strip()
                            cargo = str(row[header.get("cargo", -1)] or "").replace('\n', ' ').strip()
                            vinculo = str(row[header.get("vinculo", -1)] or "").replace('\n', ' ').strip()
                            matricula = str(row[header.get("matricula", -1)] or "").replace('\n', ' ').strip()
                            if "PAES" not in f"{vinculo} {matricula}".upper():
                                continue
                            plantoes = []
                            for dia in range(1, 32):
                                idx = header.get(dia)
                                if idx is None or idx >= len(row): continue
                                cell = row[idx]
                                if not cell: continue
                                for turno in interpretar_turno(str(cell)):
                                    data_plantao = datetime(ano, mes, dia)
                                    if turno["turno"] == "NOITE (fim)":
                                        data_plantao += timedelta(days=1)
                                    horario = HORARIOS_TURNO[turno["turno"]]
                                    plantoes.append({
                                        "dia": data_plantao.day,
                                        "data": data_plantao.strftime("%d/%m/%Y"),
                                        "turno": turno["turno"],
                                        "inicio": horario["inicio"],
                                        "fim": horario["fim"],
                                        "setor": nome_setor,
                                        "medico_setor": nome_setor
                                    })
                            if plantoes:
                                profissionais.append({
                                    "medico_nome": nome,
                                    "medico_crm": crm,
                                    "medico_especialidade": cargo,
                                    "medico_vinculo": vinculo,
                                    "medico_setor": nome_setor,
                                    "medico_unidade": nome_unidade,
                                    "plantoes": dedup_plantao(plantoes)
                                })
        return profissionais
    except Exception as e:
        print(f"Erro processando {page_info}: {str(e)}")
        return []

@app.post("/normaliza-escala-MATERNIDADE-MATRICIAL")
async def normaliza_escala_maternidade_matricial(request: Request):
    print("Requisição recebida:", await request.json())
    try:
        body = await request.json()
        todos_profissionais = []

        # Usar lista de âncora como base
        for prof in PROFISSIONAIS_ANCHOR:
            todos_profissionais.append({
                "medico_nome": prof["medico_nome"],
                "medico_crm": "",
                "medico_especialidade": "",
                "medico_vinculo": "R.P. PAES",  # Assumido como padrão
                "medico_setor": prof["medico_setor"],
                "medico_unidade": prof["medico_unidade"],
                "plantoes": []  # Plantões vazios por padrão, a serem preenchidos se PDF fornecido
            })

        # Processar entrada como array (se houver PDFs)
        if isinstance(body, list):
            for idx, item in enumerate(body):
                if "data" in item and isinstance(item["data"], list):
                    for page_data in item["data"]:
                        b64 = page_data.get("base64") or page_data.get("bae64")
                        page_number = page_data.get("page_number", "unknown")
                        if b64:
                            profs = processar_pagina_pdf(b64, f"Item {idx+1}, página {page_number}")
                            todos_profissionais.extend(profs)
                else:
                    b64 = item.get("base64") or item.get("bae64")
                    if b64:
                        profs = processar_pagina_pdf(b64, f"Item {idx+1}")
                        todos_profissionais.extend(profs)

        # Agrupar por médico para consolidar diferentes escalas
        medicos_consolidados = {}
        for prof in todos_profissionais:
            nome = prof["medico_nome"]
            if nome not in medicos_consolidados:
                medicos_consolidados[nome] = []
            medicos_consolidados[nome].append(prof)
        
        # Criar lista final mantendo escalas separadas para cada médico
        profissionais_final = []
        for nome, escalas in medicos_consolidados.items():
            if len(escalas) == 1:
                profissionais_final.append(escalas[0])
            else:
                for escala in escalas:
                    profissionais_final.append(escala)
        
        profissionais_final.sort(key=lambda p: (p["medico_nome"], p["medico_setor"]))

        # Determinar mês/ano da escala
        mes_nome_str = "JULHO"
        ano = 2025
        if profissionais_final:
            primeiro_plantao = profissionais_final[0]["plantoes"][0] if profissionais_final[0]["plantoes"] else None
            if primeiro_plantao:
                data_parts = primeiro_plantao["data"].split("/")
                mes = int(data_parts[1])
                ano = int(data_parts[2])
                mes_nome_str = [k for k, v in MONTH_MAP.items() if v == mes][0]

        return JSONResponse(content=[{
            "unidade_escala": "MISTA",
            "mes_ano_escala": f"{mes_nome_str}/{ano}",
            "profissionais": profissionais_final
        }])

    except Exception as e:
        print(f"Erro no endpoint: {str(e)}")
        return JSONResponse(content={"error": str(e)}, status_code=500)

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
# --- INICIO normaliza-MATERNIDADE-MATRICIAL ---
